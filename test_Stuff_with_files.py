import csv
import io
import os
import shutil
from io import BytesIO
from os.path import basename
from zipfile import ZipFile
import PyPDF2
from openpyxl import load_workbook

files = ('Software Testing.pdf', 'file_example_XLSX_50.xlsx', 'username.csv')
project_path = os.path.abspath('')
zip_name = 'MyZip.zip'
dir_name = 'Resources'
pdf_check = 'Тестирование программного обеспечения'
csv_check = 'booker12;9012;Rachel;Booker'


def zipping(filename='zip.zip'):
    with ZipFile(f'{filename}', mode='w') as filename:
        for file in files:
            file_path = os.path.join(project_path, file)
            filename.write(file_path, arcname=basename(file_path))


def put_file_in_dir(file):
    dir_path = os.path.join(project_path, dir_name)
    zip_path = os.path.join(project_path, dir_name, file)
    if not os.path.exists(dir_path):
        os.mkdir(os.path.join(project_path, dir_name))
    if not os.path.exists(zip_path):
        shutil.move(os.path.join(project_path, file), os.path.join(project_path, dir_name))


def check_csv():
    with ZipFile(os.path.join(project_path, dir_name, zip_name)) as zf:
        with zf.open(files[2]) as opfl:
            csv_list = csv.reader(io.TextIOWrapper(opfl, 'UTF-8'))
            for row in csv_list:
                if csv_check in row:
                    assert True
                    print(csv_check)
                    print(f'{files[2]} - имеет расширение csv\n')
                    break


def check_xlsx():
    with ZipFile(os.path.join(project_path, dir_name, zip_name)) as zf:
        with zf.open(files[1]) as opfl:
            xlsx = load_workbook(opfl).active
            try:
                print(xlsx.cell(row=3, column=2).value)
            finally:
                print(f'{files[1]} - имеет расширение xlsx\n')


def check_pdf():
    with ZipFile(os.path.join(project_path, dir_name, zip_name)) as zf:
        with zf.open(files[0]) as opfl:
            pdf = PyPDF2.PdfFileReader(BytesIO(opfl.read()))
            for page in range(0, pdf.numPages - 1):
                if pdf_check in pdf.pages[page].extract_text():
                    print(f'{pdf_check} - найден в файле')
                    print(f'{files[0]} - имеет расширения pdf\n')
                    break
            assert pdf.numPages == 303


def remove(file):
    zip_path = os.path.join(project_path, dir_name, file)
    if os.path.exists(zip_path):
        os.remove(os.path.join(project_path, dir_name, zip_name))
        print('Ресурсы очищены')


zipping(zip_name)
put_file_in_dir(zip_name)
check_csv()
check_xlsx()
check_pdf()
# remove(zip_name)
